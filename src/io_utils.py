from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable, Optional

import pandas as pd
from openpyxl import load_workbook


def clean_column_name(name: object) -> str:
    """Convert a column name to a stable snake_case-ish string while preserving Arabic text."""
    if name is None:
        return ""
    text = str(name).strip()
    text = text.replace("\ufeff", "")
    text = re.sub(r"[\s\-\/]+", "_", text)
    text = re.sub(r"[^0-9A-Za-z_\u0600-\u06FF&]+", "", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    cols = []
    seen = {}
    for c in out.columns:
        name = clean_column_name(c) or "unnamed"
        base = name
        if base in seen:
            seen[base] += 1
            name = f"{base}_{seen[base]}"
        else:
            seen[base] = 0
        cols.append(name)
    out.columns = cols
    return out


def _choose_excel_sheet(path: Path, sheet_name: Optional[str | int]) -> str | int:
    if sheet_name is not None:
        return sheet_name
    xls = pd.ExcelFile(path)
    # Prefer the real data sheet in the uploaded diversification workbook.
    for preferred in ["Data", "data", "Panel", "panel", "Sheet1"]:
        if preferred in xls.sheet_names:
            return preferred
    return xls.sheet_names[0]


def _detect_header_row(raw: pd.DataFrame) -> int:
    """Detect the header row among the first few rows of an Excel sheet.

    This handles workbooks where row 1 is a title and row 2 contains true columns,
    as in the uploaded EGX diversification file.
    """
    max_check = min(10, len(raw))
    best_idx = 0
    best_score = -1
    key_terms = {"id", "firm", "company", "year", "السنة", "اسم"}
    for i in range(max_check):
        vals = [str(x).strip().lower() for x in raw.iloc[i].tolist() if pd.notna(x) and str(x).strip()]
        nonmissing = len(vals)
        key_hits = sum(any(k in v for k in key_terms) for v in vals)
        unique_vals = len(set(vals))
        score = nonmissing + 3 * key_hits + 0.1 * unique_vals
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx



def _read_excel_values_fast(path: Path, sheet_name: str | int | None = None, max_cols: int = 200, empty_stop: int = 50) -> pd.DataFrame:
    """Read visible used values from Excel without trusting inflated worksheet dimensions."""
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name is None:
        sheet = _choose_excel_sheet(path, None)
    else:
        sheet = sheet_name
    ws = wb.worksheets[sheet] if isinstance(sheet, int) else wb[str(sheet)]
    rows = []
    consecutive_empty = 0
    seen_nonempty = False
    max_col = min(ws.max_column or max_cols, max_cols)
    for row in ws.iter_rows(min_row=1, max_col=max_col, values_only=True):
        vals = list(row)
        is_empty = all(v is None or str(v).strip() == "" for v in vals)
        if is_empty:
            if seen_nonempty:
                consecutive_empty += 1
                if consecutive_empty >= empty_stop:
                    break
            continue
        seen_nonempty = True
        consecutive_empty = 0
        # trim trailing empty cells
        while vals and (vals[-1] is None or str(vals[-1]).strip() == ""):
            vals.pop()
        rows.append(vals)
    wb.close()
    if not rows:
        return pd.DataFrame()
    width = max(len(r) for r in rows)
    rows = [r + [None] * (width - len(r)) for r in rows]
    return pd.DataFrame(rows)

def read_table(path: Path, sheet_name: Optional[str | int] = None) -> pd.DataFrame:
    """Read a CSV or Excel file with light cleaning and smart header detection."""
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt"}:
        df = pd.read_csv(path)
    elif suffix in {".xlsx", ".xls", ".xlsm"}:
        raw = _read_excel_values_fast(path, sheet_name=sheet_name)
        # Drop fully empty rows/columns caused by Excel's used-range artifacts.
        raw = raw.dropna(axis=0, how="all").dropna(axis=1, how="all")
        if raw.empty:
            return pd.DataFrame()
        header_idx = _detect_header_row(raw)
        header = raw.iloc[header_idx].tolist()
        df = raw.iloc[header_idx + 1 :].copy()
        df.columns = header
        df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    else:
        raise ValueError(f"Unsupported file type: {path}")
    return clean_columns(df)


def read_all_excel_sheets(path: Path) -> pd.DataFrame:
    """Read all sheets from an Excel workbook and concatenate them.

    Adds source_sheet for traceability. Useful for governance workbooks where each
    year is a separate sheet.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    frames = []
    for sheet in sheet_names:
        try:
            df = read_table(path, sheet_name=sheet)
            if not df.empty:
                df["source_sheet"] = sheet
                frames.append(df)
        except Exception as e:
            frames.append(pd.DataFrame({"source_sheet": [sheet], "read_error": [str(e)]}))
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True, sort=False)


def excel_inventory(path: Path) -> pd.DataFrame:
    """Return an inventory of sheets, dimensions, and columns in an Excel file."""
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    rows = []
    for sheet in sheet_names:
        df = read_table(path, sheet_name=sheet)
        rows.append(
            {
                "file": path.name,
                "sheet": sheet,
                "n_rows": int(df.shape[0]),
                "n_columns": int(df.shape[1]),
                "columns": "; ".join(map(str, df.columns)),
                "clean_columns": "; ".join(clean_column_name(c) for c in df.columns),
            }
        )
    return pd.DataFrame(rows)


def first_existing_column(df: pd.DataFrame, candidates: Iterable[str]) -> str | None:
    normalized = {clean_column_name(c).lower(): c for c in df.columns}
    for candidate in candidates:
        c = clean_column_name(candidate).lower()
        if c in normalized:
            return normalized[c]
    return None


def require_columns(df: pd.DataFrame, columns: Iterable[str], context: str = "data") -> None:
    missing = [c for c in columns if c not in df.columns]
    if missing:
        raise KeyError(f"Missing required columns in {context}: {missing}. Available columns: {list(df.columns)}")
